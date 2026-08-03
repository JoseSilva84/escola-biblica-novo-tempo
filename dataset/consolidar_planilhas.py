#!/usr/bin/env python3
"""Consolida novos Excels na planilha principal, inserindo apenas alunos ausentes."""

from __future__ import annotations

import argparse
import json
from copy import copy
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_DATASET_DIR = Path(__file__).resolve().parent
DEFAULT_BASE = DEFAULT_DATASET_DIR / "ListagemCompleta (1).xlsx"


def normalize_id(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text or None


def composite_key(row, headers):
    values = {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
    parts = [
        values.get("Aluno"),
        values.get("Sobrenome"),
        values.get("Email"),
        values.get("Telefone"),
        values.get("Data de aniversário"),
    ]
    normalized = [str(part).strip().casefold() for part in parts if part not in (None, "")]
    return "|".join(normalized) if normalized else None


def normalize_text_key(value):
    text = str(value or "").strip().casefold()
    return " ".join(text.split()) or None


def duplicate_items(counter: Counter, limit=20):
    return [
        {"valor": value, "quantidade": quantity}
        for value, quantity in counter.most_common()
        if value and quantity > 1
    ][:limit]


def copy_row_style(sheet, source_row, target_row, columns):
    for column in range(1, columns + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.border = copy(source.border)


def consolidar_planilhas(base_path: Path = DEFAULT_BASE, source_paths: list[Path] | None = None) -> dict:
    source_paths = source_paths or []
    workbook = load_workbook(base_path)
    sheet = workbook["Sheet0"]

    base_headers = [cell.value for cell in sheet[1]]
    id_col = base_headers.index("ID")
    existing_ids = set()
    existing_composites = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_id = normalize_id(row[id_col])
        if row_id:
            existing_ids.add(row_id)
        fallback = composite_key(row, base_headers)
        if fallback:
            existing_composites.add(fallback)

    original_rows = sheet.max_row - 1
    template_row = sheet.max_row
    rows_to_append = []
    new_districts = Counter()
    seen_new_ids = set()
    seen_new_composites = set()
    sources = []
    upload_ids = Counter()
    upload_emails = Counter()
    upload_names = Counter()

    for source_path in source_paths:
        source_workbook = load_workbook(source_path, read_only=True, data_only=False)
        source_sheet = source_workbook["Sheet0"]
        source_headers = [cell.value for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]
        source_id_col = source_headers.index("ID")
        source_index = {header: index for index, header in enumerate(source_headers)}
        stats = {
            "arquivo": source_path.name,
            "lidos": 0,
            "novos": 0,
            "ja_existiam": 0,
            "duplicados_upload": 0,
        }

        for row in source_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            stats["lidos"] += 1
            row_id = normalize_id(row[source_id_col])
            fallback = composite_key(row, source_headers)
            values = {header: row[index] if index < len(row) else None for index, header in enumerate(source_headers)}
            email_key = normalize_text_key(values.get("Email"))
            name_key = normalize_text_key(f"{values.get('Aluno') or ''} {values.get('Sobrenome') or ''}")
            if row_id:
                upload_ids[row_id] += 1
            if email_key:
                upload_emails[email_key] += 1
            if name_key:
                upload_names[name_key] += 1
            already_exists = (row_id and row_id in existing_ids) or (
                not row_id and fallback in existing_composites
            )
            already_seen = (row_id and row_id in seen_new_ids) or (
                not row_id and fallback in seen_new_composites
            )

            if already_exists:
                stats["ja_existiam"] += 1
                continue
            if already_seen:
                stats["duplicados_upload"] += 1
                continue

            output_row = [
                row[source_index[header]] if header in source_index else None
                for header in base_headers
            ]
            rows_to_append.append(output_row)
            district = output_row[base_headers.index("Distrito")] if "Distrito" in base_headers else ""
            new_districts[str(district or "Nao informado").strip() or "Nao informado"] += 1
            if row_id:
                seen_new_ids.add(row_id)
                existing_ids.add(row_id)
            if fallback:
                seen_new_composites.add(fallback)
                existing_composites.add(fallback)
            stats["novos"] += 1

        sources.append(stats)
        source_workbook.close()

    start_row = sheet.max_row + 1
    for offset, values in enumerate(rows_to_append):
        target_row = start_row + offset
        for column, value in enumerate(values, start=1):
            sheet.cell(row=target_row, column=column).value = value
        copy_row_style(sheet, template_row, target_row, len(base_headers))

    if rows_to_append:
        workbook.save(base_path)
    workbook.close()

    return {
        "base": str(base_path),
        "linhas_antes": original_rows,
        "alunos_novos": len(rows_to_append),
        "linhas_depois": original_rows + len(rows_to_append),
        "distritos_novos": [
            {"distrito": district, "quantidade": quantity}
            for district, quantity in sorted(new_districts.items(), key=lambda item: (-item[1], item[0]))
        ],
        "alertas_duplicidade": {
            "ids_repetidos_upload": duplicate_items(upload_ids),
            "emails_repetidos_upload": duplicate_items(upload_emails),
            "nomes_repetidos_upload": duplicate_items(upload_names),
        },
        "arquivos": sources,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Acrescenta alunos ausentes na planilha base.")
    parser.add_argument("--base", type=Path, default=DEFAULT_BASE)
    parser.add_argument("arquivos", nargs="+", type=Path)
    args = parser.parse_args()
    print(json.dumps(consolidar_planilhas(args.base, args.arquivos), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
