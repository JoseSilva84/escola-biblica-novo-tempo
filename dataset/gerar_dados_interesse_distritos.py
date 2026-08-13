#!/usr/bin/env python3
"""Gera dados_interesse_<distrito>.json para todos os distritos da planilha."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_DATASET_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = DEFAULT_DATASET_DIR / "ListagemCompleta (1).xlsx"
TARGETS = ("respondeu", "demonstrou_interesse", "aceitou_visita", "participou")
EMAIL_RE = re.compile(r"^[^@\s;]+@[^@\s;]+\.[^@\s;]+$")
MATERIAL_SPLIT_RE = re.compile(r"\s*\*-\*\s*")


def repair_mojibake(value: str) -> str:
    if not any(marker in value for marker in ("Ã", "Â", "â")):
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value


def normalize_text(value: Any) -> str:
    text = repair_mojibake(" ".join(str(value or "").strip().split()))
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char)).upper()


def display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return repair_mojibake(str(value).strip())


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = "".join(char for char in normalized if not unicodedata.combining(char))
    slug = re.sub(r"[^A-Za-z0-9]+", "_", ascii_text).strip("_")
    return slug or "Distrito"


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def material_names(value: Any) -> list[str]:
    text = display_text(value)
    if not text:
        return []
    names: list[str] = []
    for item in MATERIAL_SPLIT_RE.split(text):
        item = item.strip()
        if not item:
            continue
        name = item.split(" | ", 1)[0]
        name = re.sub(r"\s*-\s*(IMPRESSO|PDF|DIGITAL|ON-LINE|ONLINE)\s*$", "", name, flags=re.IGNORECASE)
        names.append(normalize_text(name))
    return names


def valid_phone(value: Any) -> bool:
    digits = re.sub(r"\D", "", display_text(value))
    return 10 <= len(digits) <= 13


def valid_email(value: Any) -> bool:
    emails = [part.strip() for part in re.split(r"[;,]", display_text(value)) if part.strip()]
    return bool(emails) and all(EMAIL_RE.match(email) for email in emails)


def empty_outcomes() -> dict[str, Any]:
    return {
        "tentativa_contato": False,
        "data_tentativa": None,
        "canal": None,
        "respondeu": None,
        "demonstrou_interesse": None,
        "aceitou_visita": None,
        "participou": None,
        "observacao": None,
    }


def load_existing_outcomes(path: Path) -> dict[str, dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    return {
        str(record.get("id")): record.get("resultados", {})
        for record in payload.get("registros", [])
        if record.get("id") is not None
    }


def read_rows(source: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Sheet0"]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_text(value).title() for value in next(rows)]
    canonical = {
        "Id": "ID",
        "Aluno": "Aluno",
        "Sobrenome": "Sobrenome",
        "Cidade": "Cidade",
        "Bairro": "Bairro",
        "Email": "Email",
        "Telefone": "Telefone",
        "Material": "Material",
        "Vip": "Vip",
        "Descricao": "Descricao",
        "Distrito": "Distrito",
        "Data Do Ultimo Contato": "Data do Ultimo Contato",
    }
    headers = [canonical.get(header, header) for header in headers]
    records = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        records.append({header: row[index] if index < len(row) else "" for index, header in enumerate(headers)})
    workbook.close()
    return records


def reference_date(rows: list[dict[str, Any]]) -> date:
    dates = [parsed for parsed in (parse_date(row.get("Data do Ultimo Contato")) for row in rows) if parsed]
    return max(dates) if dates else date.today()


def build_features(row: dict[str, Any], ref_date: date) -> dict[str, Any]:
    last_contact = parse_date(row.get("Data do Ultimo Contato"))
    days = max(0, (ref_date - last_contact).days) if last_contact else 365 * 20
    materials = material_names(row.get("Material"))
    phone = display_text(row.get("Telefone"))
    email = display_text(row.get("Email"))
    return {
        "log_dias_desde_contato": math.log1p(days),
        "materiais_quantidade": min(len(materials), 20),
        "tem_telefone": int(bool(phone)),
        "telefone_valido": int(valid_phone(phone)),
        "tem_email": int(bool(email)),
        "email_valido": int(valid_email(email)),
        "tem_descricao": int(bool(display_text(row.get("Descricao")))),
        "cidade": normalize_text(row.get("Cidade")) or "NAO INFORMADO",
        "bairro": normalize_text(row.get("Bairro")) or "NAO INFORMADO",
        "material_principal": materials[0] if materials else "NAO INFORMADO",
    }


def build_record(row: dict[str, Any], ref_date: date, previous: dict[str, dict[str, Any]]) -> dict[str, Any]:
    identifier = display_text(row.get("ID"))
    outcomes = empty_outcomes()
    outcomes.update(previous.get(identifier, {}))
    last_contact = parse_date(row.get("Data do Ultimo Contato"))
    return {
        "id": identifier,
        "contato": {
            "nome": " ".join(filter(None, (display_text(row.get("Aluno")), display_text(row.get("Sobrenome"))))),
            "telefone": display_text(row.get("Telefone")),
            "email": display_text(row.get("Email")),
        },
        "origem": {
            "distrito": display_text(row.get("Distrito")),
            "cidade": display_text(row.get("Cidade")),
            "bairro": display_text(row.get("Bairro")),
            "material": display_text(row.get("Material")),
            "ultimo_contato": last_contact.isoformat() if last_contact else None,
            "vip_historico": normalize_text(row.get("Vip")) == "SIM",
        },
        "atributos_modelo": build_features(row, ref_date),
        "resultados": outcomes,
    }


def gerar_dados_interesse_distritos(source: Path = DEFAULT_EXCEL, output_dir: Path = DEFAULT_DATASET_DIR) -> dict[str, Any]:
    rows = read_rows(source)
    ref_date = reference_date(rows)
    by_district: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        district = display_text(row.get("Distrito")) or "Nao informado"
        by_district.setdefault(district, []).append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    files = []
    for district, district_rows in sorted(by_district.items(), key=lambda item: normalize_text(item[0])):
        output = output_dir / f"dados_interesse_{slugify(district)}.json"
        previous = load_existing_outcomes(output)
        records = [build_record(row, ref_date, previous) for row in district_rows]
        payload = {
            "schema_version": 1,
            "objetivo": "Prever resultado real de contato, nao reproduzir o VIP historico.",
            "distrito_piloto": district,
            "fonte": str(source),
            "data_referencia": ref_date.isoformat(),
            "total_registros": len(records),
            "alvos_disponiveis": list(TARGETS),
            "campos_excluidos_modelo": [
                "nome",
                "telefone",
                "email",
                "sexo",
                "religiao",
                "idade",
                "endereco",
                "vip_historico",
                "distrito",
                "cidade",
                "bairro",
            ],
            "registros": records,
        }
        output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        files.append({"distrito": district, "arquivo": output.name, "registros": len(records)})

    summary = {"total_distritos": len(files), "total_registros": len(rows), "arquivos": files}
    (output_dir / "dados_interesse_distritos_manifest.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera dados_interesse_<distrito>.json para todos os distritos.")
    parser.add_argument("--arquivo", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--saida", type=Path, default=DEFAULT_DATASET_DIR)
    args = parser.parse_args()
    print(json.dumps(gerar_dados_interesse_distritos(args.arquivo, args.saida), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
