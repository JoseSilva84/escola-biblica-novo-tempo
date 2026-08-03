#!/usr/bin/env python3
"""Gera alunos.json a partir da planilha principal do dataset."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_DATASET_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = DEFAULT_DATASET_DIR / "ListagemCompleta (1).xlsx"
DEFAULT_OUTPUT = DEFAULT_DATASET_DIR / "alunos.json"


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def gerar_alunos_json(source: Path = DEFAULT_EXCEL, output: Path = DEFAULT_OUTPUT) -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["Sheet0"]

    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]

    records = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        records.append(
            {
                header: normalize_value(row[index] if index < len(row) else "")
                for index, header in enumerate(headers)
            }
        )

    workbook.close()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return len(records)


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera dataset/alunos.json a partir do Excel.")
    parser.add_argument("--arquivo", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--saida", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    total = gerar_alunos_json(args.arquivo, args.saida)
    print(f"alunos.json gerado: {total} registros")


if __name__ == "__main__":
    main()
